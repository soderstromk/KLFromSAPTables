from tkinter import *
from tkinter import ttk
import pymysql
import os
import numpy

class KLGUI: #This is the GUI window that accepts input for KL distance measures
    def __init__(self, master, Database): #Database = sap the SAP 2011 default MySQL database
        #These are the elements of the GUI
        mainframe = Frame(master,width=150,height=150)
        mainframe.pack()

        #Output directory control
        fileframe = Frame(mainframe,width=150, height=10)
        fileframe.pack(anchor="w",expand="true")

        self.FileLabel = Label(fileframe,text='Directory to Store Results')
        self.FileLabel.grid(column=0,row=0,sticky=W)
        self.SaveDir = StringVar()
        path = os.path.normpath("c:/sap/kl_results/") #Default path to store results
        self.SaveDir.set(path)
        self.FileInput = Entry(fileframe,width=60)
        self.FileInput["textvariable"]= self.SaveDir
        self.FileInput.grid(column=0, row=1,sticky=W)
        self.FileButton = Button(fileframe,text='Change Directory',command=self.ChangeDir)
        self.FileButton.grid(column=1, row=1,sticky=E)

        #Code for Template ListBox

        templateframe = Frame(mainframe,width = 70, height=20)
        templateframe.pack(side=LEFT)

        #Tempsliderframe = Frame(mainframe,width=70, height=5)
        #Tempsliderframe.pack(side=LEFT)

        targetframe=Frame(mainframe,width=70,height=20)
        targetframe.pack(side=LEFT)

        self.Targetsliderframe = Frame(mainframe,width=70,height=20)
        self.Targetsliderframe.pack(side=LEFT)

        self.TemplateText = StringVar()
        self.TemplateText.set('Select Template Table')
        self.TemplateListLabel = Label(templateframe,textvariable = self.TemplateText,anchor="w")
        self.TemplateListLabel.pack()

        self.TemplateListBox = Listbox(templateframe, selectmode=BROWSE)
        self.TemplateListBox.pack(side=LEFT, fill=Y)

        self.TemplateListScroll = Scrollbar(templateframe)
        self.TemplateListScroll.pack(side=LEFT, fill=Y)

        sylselectframe = Frame(mainframe,width=20,height=20)
        sylselectframe.pack()

        self.TemplateListScroll.config(command=self.TemplateListBox.yview)
        self.TemplateListBox.config(yscrollcommand=self.TemplateListScroll.set)

        self.TableValues = []
        #GetTableNames generates a list of tables in the sap MySQL database
        Tables = self.GetTableNames(Database)

        #Adds table names to list box where they can be selected in the GUI
        self.ListOfTableNames = []
        self.Sliders ={}
        self.SylDur={}


        self.TemplateSlider=Scale(self.Targetsliderframe)
        self.savemotifs = BooleanVar()
        self.motifsavebutton=Checkbutton(self.Targetsliderframe,text='Save syllable duration files?',variable=self.savemotifs,onvalue=TRUE,offvalue=NONE)

        for name in Tables:
            self.ListOfTableNames.append(name)
            self.TemplateListBox.insert(END,name)

            #make a syllable interval slider for every table just in case
            sylslide=Scale(self.Targetsliderframe,from_=25, to=100, orient=HORIZONTAL,
                                                            label=name[0][5:] + ' Motif Interval Cutoff (ms)',
                                                            sliderlength=10, length=250)
            self.Sliders[name[0]] = sylslide #add the slider to the dictionary

            if not name[0] in self.SylDur.keys():
                self.SylDur[name[0]] = 35 #if it hasn't been set use default syldur = 35 ms
            self.Sliders[name[0]].set(self.SylDur[name[0]])

        self.TemplateListBox.bind('<<ListboxSelect>>', self.TemplateListBoxChanged)
        self.TemplateName = StringVar()

        #Adds table names to list box where they can be selected in the GUI
        #Code for Target Table ListBox
        self.TargetText = StringVar()
        self.TargetText.set('Select Target Table(s)')
        self.TargetListLabel = Label(targetframe, textvariable=self.TargetText,anchor="w")
        self.TargetListLabel.pack()

        # List box accesses SAP database and lists available MySQL tables created by SAP 2011
        self.TargetListBox = Listbox(targetframe, selectmode=MULTIPLE)
        self.TargetListScroll = Scrollbar(targetframe)
        self.TargetListBox.config(yscrollcommand=self.TargetListScroll.set)
        self.TargetListScroll.config(command=self.TargetListBox.yview)

        #Make the target list box
        for i in Tables:
            self.TargetListBox.insert(END, i[0])

        self.TargetListBox.pack(side=LEFT, fill=Y)
        self.TargetListBox.bind('<<ListboxSelect>>', self.TargetListBoxChanged)
        #self.TargetListScroll.grid(row=5,column=1)
        self.TargetListScroll.pack(side=RIGHT, fill=Y)
        #self.TargetListBox.grid(column=0,row=5)

        self.TargetTableList = []

        #GUI to select whether motif vs non-motif are determined and if this is used in KL distance measures
        self.motif=IntVar()
        self.Radio_1 = Radiobutton(sylselectframe, text="Use All Syllables", variable=self.motif, value=0,command=self.MotifChange) #Does not distinguish Motif vs. non-Motif syllables
        self.Radio_2 = Radiobutton(sylselectframe, text="Filter Out Non-Motif Syllables", variable=self.motif,value=1,command=self.MotifChange) #Only uses Motif syllables for KL calculation
        self.Radio_3 = Radiobutton(sylselectframe, text="Calculate Motif-Type but Use All", variable=self.motif,value=2,command=self.MotifChange) #Determines Motif vs. non-Motif syllables, but uses all for KL
        self.Radio_1.pack(side=TOP,anchor='w')
        self.Radio_2.pack(side=TOP,anchor='w')
        self.Radio_3.pack(side=TOP,anchor='w')

        self.AxesDivisions = IntVar() #Determines the number of bins that each axis is divided into
        self.AxesDivisions.set(15) #Default bins = 15. Given 2 axes the number of cells = 15^2 = 225
        self.AxesScale = Scale(sylselectframe, from_=5, to=50, orient = HORIZONTAL, label = 'Set Axes Divisions', sliderlength = 10, length = 200, variable=self.AxesDivisions)
        self.AxesScale["variable"] = self.AxesDivisions
        self.AxesScale.pack(side=TOP)

        # If a *.wav file contains fewer than this minimum number of syllables it is not used for KL calculation
        # GUI code to select minimum syllable number for a *.wav to be used
        self.MinSyllPerWav = IntVar()
        self.MinSyllPerWav.set(5)
        self.MinSyllScale = Scale(sylselectframe, from_=1, to=10, orient=HORIZONTAL,
                                  label='Set Minimum Syllables/WAV File', sliderlength=10, length=200,
                                  variable=self.MinSyllPerWav)
        self.MinSyllScale.pack(side=TOP)

        # Smoothing factors are used to replace cells with no syllables as the KL distance equation cannot use 0 values
        self.SmoothingFactor = DoubleVar()
        self.SmoothingHolder = IntVar()
        self.SmoothingHolder.set(-6)
        self.SmoothingScale = Scale(sylselectframe, from_=-12, to=0, command=self.CalcSmoothFactor, orient=HORIZONTAL,
                                    label='Set KL Smoothing Factor 10^x', sliderlength=10, length=200,
                                    variable=self.SmoothingHolder)
        self.SmoothingScale.pack(side=TOP)

        #GUI to launch KL distance analysis
        self.KLButton = Button(sylselectframe,text='Run KL',command=self.RunKL,anchor=SE)

        self.KLButton.pack(side=RIGHT, padx=(10, 0))#Pack after MotifInterval to prevent overlap

        self.StatusLabel=Label(fileframe,text='')
        self.StatusLabel.grid(column=0,row=2)

    #Function to transform the smooting factor input to a log scale. The function that generates the probability table needs log values
    def CalcSmoothFactor(self, evt):
        self.SmoothingFactor.set(10**self.SmoothingHolder.get())

    # Code to assign a new output path to save Excel file results in if it is changed in the GUI
    def ChangeDir(self):
        self.StatusLabel.pack_forget()
        from tkinter import filedialog
        path = filedialog.askdirectory()
        self.FileInput.delete(0, END)
        if path=='':
            path='c:/sap/kl_results/'
        self.FileInput.insert(0, path)

    #Function to extract names of tables from the sap MySQL database so that one can be selected for analysis
    def GetTableNames(self, db):
        conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap')
        cursor = conn.cursor()
        query = "USE " + db
        cursor.execute(query)
        query = """SHOW TABLES LIKE %s"""
        cursor.execute(query, (('%' + 'syll' + '%',)))
        tables = cursor.fetchall()
        conn.close()
        cursor.close()
        return tables

    #Function to change the template table selected
    def TemplateListBoxChanged(self, evt):
        self.TemplateSlider.pack_forget()
        #self.TemplateSlider = Scale(self.thisframe)
        self.StatusLabel.pack_forget()
        self.ListOfDays = []
        w = evt.widget
        if w.curselection:
            idx = int(w.curselection()[0])
            now = w.get(idx)
            self.TemplateName = str(now[0])
            self.TemplateText.set("Template: " + self.TemplateName[5:])

        tempdur=self.Sliders[self.TemplateName].get()
        self.TemplateSlider = self.Sliders[self.TemplateName]
        self.TemplateSlider.set(tempdur)

        self.MotifChange()

    #For when the Target Table ListBox changes
    def TargetListBoxChanged(self,evt):
        self.StatusLabel.pack_forget()
        for table in self.TargetTableList:
            self.Sliders[table].pack_forget()
        self.TargetTableList=[]
        tablestr =''
        OldSylIntDict={}
        w = evt.widget
        idx=[]
        for i in w.curselection():
           idx.append(int(i))

        for i in idx:
            self.TargetTableList.append(w.get(i))

        #Generate text for GUI label informing user of target tables selected
        for i in self.TargetTableList:
            tablestr = tablestr + i[5:] + ", "
        tablestr=tablestr[:-2]

        self.TargetText.set("Target Table(s): "+ tablestr)

        self.MotifChange()

    #Code that shows or hides syllable interval selection on the GUI depending on if it is used for KL distance measures
    #@property

    def MotifChange(self):
        if self.TemplateName and self.TargetTableList:
            button = self.motif.get()
            if button > 0:
                #self.Targetsliderframe.pack()
                self.TemplateSlider.pack()
                self.motifsavebutton.pack()
                for counter, tablename in enumerate(self.TargetTableList):
                    if not tablename == self.TemplateName:
                        self.Sliders[tablename].pack()
            else:
                self.TemplateSlider.pack_forget()
                self.motifsavebutton.pack_forget()
                #self.Targetsliderframe.pack_forget()
                for tablename in self.TargetTableList:
                    if not tablename == self.TemplateName:
                        self.Sliders[tablename].pack_forget()

    #Creates a dictionary of values taken from GUI input. This dictionary is used by the mainKL procedure launched by the RunKL button on the GUI
    def BundleVars(self):
        tempsyldur = self.Sliders[self.TemplateName].get()
        targetsyldur={}

        for targetname in self.TargetTableList:
            targetsyldur[targetname]=self.Sliders[targetname].get()

        vars = {'FileDirectory': os.path.normpath(self.SaveDir.get()), 'TemplateToUse': self.TemplateName, 'TempSylDur':tempsyldur,
                'TargetTableList': self.TargetTableList,'TargSylDurDict': targetsyldur,
                'AxesDivisions': self.AxesDivisions.get(), \
                'OnlyMotifs': self.motif.get(), \
                'MinSyllPerWav': self.MinSyllPerWav.get(), 'SmoothingFactor': self.SmoothingFactor.get(),\
                'SaveMotifs': self.savemotifs.get()
                }
        return vars

    #The main procedure launched by the RunKL GUI button
    def RunKL(self):
        #Notification to user that the application is running
        self.StatusLabel.config(text='Running...')
        self.StatusLabel.grid(row=2,column=0)
        self.StatusLabel.update()

        #The main procedure that caluclates KL distances between template days and each target day
        Vars = self.BundleVars()#Creates the BundleVars dictionary to pass to mainKL
        mainKL(Vars)
        #Notification to user that the analysis is complete
        self.StatusLabel.config(text='Done',fg="red")

#Main function
def mainKL(Inputvars):
    #No table selected in GUI
    if Inputvars.get('TemplateToUse') == None:
        from tkinter import messagebox
        entertable = messagebox.showinfo("Template Table Not Selected", "Select Template Table")
        return
    if Inputvars.get('TempSylDur') == None:
        from tkinter import messagebox
        entertable = messagebox.showinfo("Target Table Not Selected", "Select Target Table")
        return

    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap')
    cur = conn.cursor()

    query = "drop table if exists target"
    cur.execute(query)

    query = "drop table if exists template"
    cur.execute(query)

    query = "drop table if exists "+ Inputvars.get('TemplateToUse')[5:] +"_syldur"
    cur.execute(query)

    for Table in Inputvars.get('TargetTableList'):
        query = "drop table if exists " + Table[5:] + "_syldur"
        cur.execute(query)

    if Inputvars.get('OnlyMotifs'): #Indicated in GUI, could be 1 = Only Motifs or 2 = Use All And Calc Syl Type
        #Calls function to calculate
        Syll_Duration(Inputvars.get("TemplateToUse"), Inputvars.get('TempSylDur'),
                  Inputvars.get("MinSyllPerWav"))  # Adds pre/post interval and syllable type(motif v non-motif) 'templatename_SYLDUR'

        for Target in Inputvars.get('TargetTableList'):
            Syll_Duration(Target, Inputvars.get('TargSylDurDict')[Target],Inputvars.get("MinSyllPerWav"))

            # Adds pre/post interval and syllable type(motif v non-motif)
            # Calls a function to create a new MySQL table that is a subset of the sap table selected by the GUI...
            # ...that includes syllable interval data 'targetname_SYLDUR'

    KLDict={}
    MakeTemplateOrTargetTable(Inputvars.get('TemplateToUse'), 'TEMPLATE', Inputvars.get("OnlyMotifs", None)) # Creates table 'TEMPLATE'

    for TargetTable in Inputvars.get('TargetTableList'):
        MakeTemplateOrTargetTable(TargetTable, 'TARGET', Inputvars.get("OnlyMotifs", None))  # Creates table 'TARGET'

        #A list of SAP 2011 parameters for convenience
        paramlist = ['duration', 'mean_amplitude', 'mean_pitch', 'mean_FM', 'mean_am2', 'mean_entropy',
                 'mean_goodness_of_pitch', 'mean_mean_frequency', 'var_pitch', 'var_FM', 'var_entropy',
                 'var_goodness_of_pitch', 'var_mean_frequency', 'var_am']

        XTxt = paramlist[0]  # 'duration'
        TempX = TableParameters('TEMPLATE',XTxt)#Creates a dictionary of values for Duration (always the x-axis parameter) from the Template MySQL table
        NoticeText = "Calculating KL..."
        window.StatusLabel.config(text=NoticeText)
        window.StatusLabel.update()
        #A dictionary to collect KL values for each acoustic parameter except Syllable Duration (that is always the x-axis
        KLPerParam = {}

        GotTargetX = False #Boolen switch so that TargetX only runs TableParameters once as x-axis is always Syllable Duration

        #For each of the acoustic parameters...

        for YTxt in paramlist:
        #Except for duration that is always the x-axis
            if not YTxt == 'duration': #If the x-axis 'duration' table has not been created yet
                if not GotTargetX:
                    TargetX = TableParameters('TARGET', XTxt)
                    GotTargetX = True
                    #Make the y-axis table for current YTxt from paramlist
                    #Creates a dictionary of values from the YTxt acoustic feature...
                    #...From the Template table
                TempY = TableParameters('TEMPLATE',YTxt)
                    #...From the current target day (recday) table
                TargetY = TableParameters('TARGET',YTxt)
                    #Creates a GenerateKL object for the current acoustic parameter between Template and TargetDay (recday) tables
                KL = GenerateKL(TempX[XTxt], TempY[YTxt], TargetX[XTxt], TargetY[YTxt], Inputvars.get("AxesDivisions"), Inputvars.get("SmoothingFactor"))
                #Gets the KL distance measure for the current acoustic parameter between Template and Target tables from the GenerateKL object
                KLPerParam[YTxt] = KL.Get_KLDistance()

        KLDict[TargetTable]=KLPerParam

        # Adds the KL Distance measure results for the current acoustic parameter and day to the KLDict dictionary
        # Loops to the next acoustic parameter for the Target day
        # Unless it has completed the last in which case it goes to the next Target day and starts with 'mean_amplitude'

        # KL distances for all parameters for all Target days have been calculated
        # Now save KL distance data to an Excel spreadsheet in the FileDirectory specified by the GUI

    NoticeText = "Starting to write to Excel..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    WriteXL(Inputvars.get('FileDirectory'), KLDict, Inputvars.get('TemplateToUse'), Inputvars.get('TempSylDur'), Inputvars.get('TargetTableList'),
                Inputvars.get('TargSylDurDict'), Inputvars.get('AxesDivisions'), Inputvars.get('MinSyllPerWav'), Inputvars.get("OnlyMotifs"), Inputvars.get("SmoothingFactor"))

    CleanUpTables(Inputvars.get('TemplateToUse'), Inputvars.get('TargetTableList'), Inputvars.get('OnlyMotifs'),Inputvars.get('SaveMotifs'))

def Syll_Duration(table_name, sylldurcrit, minsylls):

    SerNums = GetSerNums(table_name) #Returns list of [serial_number, syllable count] that correspond to single wav files
    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap')
    cursor = conn.cursor()

    NoticeText = "Parsing Motif vs. Non-Motif syllables..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    query = "DROP TABLE IF EXISTS " + table_name[5:] + "_SYLDUR"
    cursor.execute(query)

    query = "SELECT COUNT(1) indexExists FROM INFORMATION_SCHEMA.STATISTICS WHERE table_schema=DATABASE() AND table_name='" \
             + table_name[5:] + "_SYLDUR' AND index_name = 'serial_number'"

    cursor.execute(query)

    query = "create index serial_number ON " + table_name + " (serial_number)"
    cursor.execute(query)

    query = "CREATE TABLE " + table_name[5:] +"_SYLDUR(recnum INT(11) PRIMARY KEY, preint FLOAT, postint FLOAT, syll_type BIT)"
    cursor.execute(query)
    linenum =1
    for line in SerNums: #serial_num, Number of syllables
        NoticeText = "Parsing Motif vs. Non-Motif syllables file " + str(linenum) + " of " + str(len(SerNums))
        window.StatusLabel.config(text=NoticeText)
        window.StatusLabel.update()
        query = "SELECT recnum, serial_number, start_on, duration FROM " + table_name + " WHERE serial_number = " + str(line[0]) + " ORDER BY recnum"
        cursor.execute(query)
        wavres =[X for X in cursor.fetchall()]
        #recnum =0, serial_number = 1, start_on = 2, duration = 3             NOT ANYMORE             preint = 5, postint = 6, syll_type = 7
        syldurvals =[]

        if len(wavres) > 2: #Some files have 1 syllable. Throws an error.
            for i in range(0,len(wavres)):
                my_recnum = wavres[i][0]
                if i == 0 or i == len(wavres)-1:
                    if i == 0:
                        my_postint = wavres[i+1][2] - (wavres[i][2]+wavres[i][3])
                        if my_postint < sylldurcrit and line[1] >= minsylls:
                            my_syll_type = 1
                        else:
                            my_syll_type = 0
                        syldurvals.append((my_recnum,None,my_postint,my_syll_type))
                    else:
                        my_preint = wavres[i][2] - (wavres[i-1][2]+wavres[i-1][3])
                        if my_preint < sylldurcrit:
                            my_syll_type = 1
                        else:
                            my_syll_type = 0
                        syldurvals.append((my_recnum,my_preint,None,my_syll_type))
                else:
                    my_postint = wavres[i + 1][2] - (wavres[i][2] + wavres[i][3])
                    my_preint = wavres[i][2] - (wavres[i - 1][2] + wavres[i - 1][3])
                    if (my_preint < sylldurcrit or my_postint < sylldurcrit) and line[1] >= minsylls:
                        my_syll_type = 1
                    else:
                        my_syll_type = 0
                    syldurvals.append((my_recnum,my_preint,my_postint,my_syll_type))

            query = "INSERT INTO " + table_name[5:] + "_SYLDUR (recnum, preint, postint, syll_type) VALUES (%s, %s, %s, %s)"
            cursor.executemany(query,syldurvals)
            conn.commit()
            linenum = linenum + 1
    #table_name_SYLDUR will not be in recnum order
    query = "drop index serial_number ON " + table_name
    cursor.execute(query)
    cursor.close()
    conn.close()

#To be used as a Template by the GUI
#Depending upon GUI entry for motif types to be used for analysis (OnlyMotifs) it will modify the Template MySQL table to...
#include the syllable type column or not
def MakeTemplateOrTargetTable(TableName, TempOrTarg, OnlyMotifs):

    NoticeText = "Making " + TempOrTarg +" Table..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap')
    cursor = conn.cursor()

    query = "DROP TABLE IF EXISTS " + TempOrTarg
    cursor.execute(query)

    if TableName:
        query = "DROP TABLE IF EXISTS TEMPORARY"
        cursor.execute(query)

        query = "CREATE TABLE TEMPORARY SELECT * FROM " + TableName + " ORDER BY recnum"
        cursor.execute(query)

        query = "SELECT * FROM " + TableName + " LIMIT 1"
        cursor.execute(query)

        field_names = [i[0] for i in cursor.description]  # get list of field names to know if preint, postint, syll_type need to be added
        if 'preint' in field_names:
           query = "ALTER TABLE " + TableName + " DROP preint, DROP postint, DROP syll_type"
           cursor.execute(query)

           query = "ALTER TABLE TEMPORARY DROP preint, DROP postint, DROP syll_type"
           cursor.execute(query)

        if OnlyMotifs == 1:
            query = "ALTER TABLE TEMPORARY ADD(preint FLOAT, postint FLOAT, syll_type BIT)"
            cursor.execute(query)

            query = "UPDATE TEMPORARY AS t1 INNER JOIN " + TableName[5:] + "_SYLDUR AS t2 ON t1.recnum = t2.recnum SET t1.preint = t2.preint, t1.postint = t2.postint, t1.syll_type = t2.syll_type"
            cursor.execute(query)
            conn.commit()

            query = "CREATE TABLE "+ TempOrTarg +" SELECT * FROM TEMPORARY WHERE syll_type = 1 ORDER BY recnum"
            cursor.execute(query)
        else:
            query = "CREATE TABLE " + TempOrTarg + " SELECT * FROM TEMPORARY ORDER BY recnum"
            cursor.execute(query)

        query = "alter table " + TempOrTarg + " add primary key (recnum)"
        cursor.execute(query)

        cursor.close()
        conn.close()
    else:
        x = 'error'
    return

def CleanUpTables(TemplateID, TargetID,onlymotifs,savemotifs):
    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap')
    cursor = conn.cursor()

    #import tkinter.messagebox
    if onlymotifs > 0 and not savemotifs:
            #tkinter.messagebox.askyesno("Save Syl_Dur Tables?", "Save Motif Tables?"):
        query = "DROP TABLE IF EXISTS " + TemplateID[5:] + "_SYLDUR"
        cursor.execute(query)

        for Table in TargetID:
            query = "DROP TABLE IF EXISTS " + Table[5:] + "_SYLDUR"
            cursor.execute(query)

    query = "DROP TABLE IF EXISTS TEMPLATE"
    cursor.execute(query)

    query = "DROP TABLE IF EXISTS TARGET"
    cursor.execute(query)

    return

# Creates a dictionary of values for a particular acoustic parameter from a particular MySQL table
def TableParameters(tablename,parameter):
    ParamRes = []
    ParamDict ={}

    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap')
    cur = conn.cursor()

    query = "create index " + parameter +" ON " + tablename + " (" + parameter +")"
    cur.execute(query)

    query = "create index recnum ON " + tablename + " (recnum)"
    cur.execute(query)

    query = "SELECT recnum, " + parameter + " FROM " + tablename + " ORDER BY recnum"
    cur.execute(query)

    for i in range(0,cur.rowcount):
        res = cur.fetchone()
        ParamRes.append(res[1])

    ParamDict[parameter] = ParamRes

    query = "drop index " + parameter +" ON " + tablename
    cur.execute(query)

    query = "drop index recnum ON " + tablename
    cur.execute(query)

    cur.close()
    conn.close()

    return ParamDict

#Function to create an Excel spread sheet and write KL distance measure results
#WriteXL(Inputvars.get('FileDirectory'), KLDict, Inputvars.get('TemplateToUse'), TargetList
#           , Inputvars.get('AxesDivisions'), Inputvars.get('MinSyllPerWav'), Inputvars.get("OnlyMotifs"), Inputvars.get("SyllInterval"), Inputvars.get("SmoothingFactor"))
def WriteXL(Path, KLResults, TemplateID, TempSylDur, TargetList, TargetDict, axdiv, minsylperwav, syltype, smoothingfactor):

    NoticeText = "Write to Excel Step 1..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    import os
    import openpyxl as xl

    NoticeText = "Write to Excel Step 2..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    StartDir = os.path.normpath(Path)
    KLDir = "kl_results"
    if StartDir.find(KLDir) < 1:
        KLDir = "\kl_results"
        os.makedirs(Path + KLDir, exist_ok=True)
        StartDir = Path + KLDir
    else:
        os.makedirs(StartDir, exist_ok=True)
    os.chdir(StartDir)

    NoticeText = "Write to Excel Step 3..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    wb = xl.Workbook()
    ws = wb.active

    NoticeText = "Write to Excel Step 4..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    ws.title = "KL Results Summary"
    ws['A1'] = 'Template = ' + TemplateID[5:]
    #ws['A1'] = 'Target = ' + TargetID[5:]

    NoticeText = "Write to Excel Step 5..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    ParameterList = ['mean_amplitude', 'mean_pitch', 'mean_FM', 'mean_am2', 'mean_entropy', 'mean_goodness_of_pitch',
                     'mean_mean_frequency', 'var_pitch', 'var_FM', 'var_entropy', 'var_goodness_of_pitch',
                     'var_mean_frequency', 'var_am']
    ParameterDict = {'mean_amplitude': 1, 'mean_pitch': 2, 'mean_FM': 3, 'mean_am2': 4, 'mean_entropy': 5,
                     'mean_goodness_of_pitch': 6, 'mean_mean_frequency': 7, 'var_pitch': 8, 'var_FM': 9,
                     'var_entropy': 10, 'var_goodness_of_pitch': 11, 'var_mean_frequency': 12, 'var_am': 13}
    #col = 3

    NoticeText = "Write to Excel Step 6..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    if syltype == 1:
        syltext = 'only motifs, Template = ' + str(TempSylDur) + ' ms syl int'
        syltext=syltext[:-1]
    else:
        syltext = 'all syllables'

    ValString = str(axdiv) + " axes divisions, " + str(smoothingfactor) + " smoothing factor, " + str(
        minsylperwav) + " syllables/wav minimum , " + syltext

    ws["C1"] = ValString
    ws["P4"] = 'Mean KL'
    ws["Q4"] = "Total Syllables"
    if syltype >= 1:
        ws["R4"] = "Motif Syllables"
    else: ws["R4"] = ""
    ws["A4"] = "Target"
    ws["B4"] = "syl int (ms)"

    TargetSyll = {}
    for Table in TargetList:
        TargetSyll[Table] = SyllableNumbers(Table,syltype)

    currow = 5

    for Table in TargetList:
        ws.cell(row=currow,column=1,value=Table[5:])
        ws.cell(row=currow, column=2, value=TargetDict[Table])
        ws.cell(row=currow, column=17, value=TargetSyll[Table][0])
        if syltype >= 1:
            ws.cell(row=currow, column=18, value=TargetSyll[Table][1])
        currow = currow + 1

    for row, Table in enumerate(TargetList):
        for col, Param in enumerate(ParameterList):
            ws.cell(row=row+5, column=col+3, value=KLResults[Table][Param])

    for col, Param in enumerate(ParameterList):
        ws.cell(row=4, column=col + 3, value=Param)

    for row, Table in enumerate(TargetList):
        start = "C" + str(row + 5)
        finish = "O" + str(row + 5)
        equation = "=AVERAGE(" + start + ":" + finish + ")"
        ws.cell(row=row + 5, column=16, value=equation)

    FileName = TemplateID[5:] + "_"
    for Table in TargetList:
        FileName = FileName + Table + ", "
    FileName = FileName + SyllableText(syltype)
    FileName = FileName + " KL Results.xlsx"

    try:
        wb.save(FileName)
    except PermissionError:
        from tkinter import messagebox
        messagebox.showinfo("Before closing this message MAKE SURE to close " + FileName, FileName +" Already Open")
    return

def GetSerNums(table_name):
    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap',cursorclass=pymysql.cursors.SSCursor)
    cursor = conn.cursor()

    NoticeText = "Grouping Recording Days..."
    window.StatusLabel.config(text=NoticeText)
    window.StatusLabel.update()

    query = "DROP TABLE IF EXISTS SERNUMS"
    cursor.execute(query)

    query = "CREATE TABLE SERNUMS SELECT serial_number, COUNT(1) AS SerCount FROM " + table_name + " GROUP BY serial_number"
    cursor.execute(query)

    query = "SELECT * FROM SERNUMS ORDER BY serial_number"
    cursor.execute(query)

    res = cursor.fetchall()
    #SerNums = [[i[0], i[1]] for i in res]
    SerNums = []
    for row in res:
        #print(row)
        SerNums.append([row[0],row[1]])

    cursor.close()
    conn.close()

    return SerNums

def SyllableText(syltype):
    syltext = {
        0: " All Sylls",
        1: " Only Motifs",
        2: " Motifs Calc'd All Used",
    }
    return syltext.get(syltype,"nothing")

# Encapsulated class that generates probability tables for Template and a Target tables and uses
# these 2D arrays for KL distance measures for an acoustic parameter
class GenerateKL:
    def __init__(self, TempX, TempY, TargetX, TargetY, axediv, SmoothFactor):
        # Establishes min and max values for Template and Target
        # These are used to define the range of the axes values that will be divided into
        # axediv number of bins
        # Axes bins will be used to generate a 2D array of cells = axediv^2
        # For default axediv = 15, the number of 2D array cells = 225
        maxy = max(max(TempY),max(TargetY))
        miny = min(min(TempY),min(TargetY))
        maxx = max(max(TempX),max(TargetX))
        minx = min(min(TempX),min(TargetX))

        # The numpy procedure histogram2d is used to generate Template and Target incidence
        # distributions
        # Each incidence represents a syllable that falls into a cell of the 2D array
        self.TemplateArray, TempXEdge, TempYEdge = numpy.histogram2d(TempX, TempY, bins = \
            axediv, range = [[minx, maxx], [miny, maxy]],normed=False)  # axediv x axediv array
        #  w/counts of sylls
        self.TargetArray, TargetXEdge, TargetYEdge = numpy.histogram2d(TargetX, TargetY, \
            axediv,range = [[minx, maxx], [miny, maxy]],normed=False)

        # To each cell of the 2D Template and Target incidence distributions is added
        # a smoothing factor.
        # The magnitude of the smoothing factor can be controlled by the user through the GUI.
        # Default = 10^-6
        # A smoothing factor is necessary as the KL distance equation uses probability
        # distribution factors as denominators
        self.TemplateArray = [(x + SmoothFactor) for x in self.Get_TemplateArray()]
        self.TargetArray = [(x + SmoothFactor) for x in self.Get_TargetArray()]

        # Incidence distributions are converted to probability distributions where the sum
        # of all cells of the 2D distribution = 1
        self.TemplateProbs = [x / sum(sum(self.Get_TemplateArray())) for x in \
                              self.Get_TemplateArray()]
        self.TargetProbs = [x / sum(sum(self.Get_TargetArray())) for x in \
                            self.Get_TargetArray()]

        #Probability distributions are then passed to a function that calculates KL distance
        self.KL = self.KLCalc(self.TemplateProbs, self.TargetProbs, axediv)

    # Get function passes the TemplateArray from the GenerateKL object
    def Get_TemplateArray(self):
        return self.TemplateArray

    # Get function passes the TargetArray from the GenerateKL object
    def Get_TargetArray(self):
        return self.TargetArray

    # Get function passes the KL distance measure from the GenerateKL object
    def Get_KLDistance(self):
        return self.KL

    # This function uses Template and Target arrays to calculate KL distance
    # This gives results identical to the scipy.stats.entropy function
    # This measure was first applied to evaluating differences in acoustic features by...
    # Wu W, Thompson JA, Bertram R, Johnson F. J Neurosci Methods. 2008 Sep 15;174(1):147-54
    # http://www.sciencedirect.com/science/article/pii/S0165027008003907
    def KLCalc(self, TemplateArray, TargetArray, AxesDiv):
        import math
        SumMN = 0
        for i in range(AxesDiv):
            for j in range(AxesDiv):
                q_1 = TemplateArray[i][j]
                q_k = TargetArray[i][j]
                qratio = q_1 / q_k
                if qratio == 0:
                    SumMN = SumMN
                else:
                    SumMN = SumMN + (q_1 * math.log(qratio, 2))

        KLDistance = SumMN
        return KLDistance

#Function to count the number of syllables used for KL distance measures for each Target day
def SyllableNumbers(Table, OnlyMotifs):
    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='sap2011', db='sap')
    cursor = conn.cursor()
    motifsyllables = 0
    SyllableTup = ()

    query = "SELECT * from " + Table
    cursor.execute(query)
    totalsyllables = cursor.rowcount
    motifsyllables = 0
    if OnlyMotifs >= 1:
        query = "SELECT recnum from " + Table[5:] +"_syldur WHERE syll_type = 1"
        cursor.execute(query)
        motifsyllables = cursor.rowcount

    SyllableTup = (totalsyllables, motifsyllables)

    return SyllableTup

#This is a Tkinter GUI program
root = Tk()
root.title("KLFromSAPTables")
#Calls the GUI
window = KLGUI(root, Database='sap')
#Infinite loop while the GUI is active
root.mainloop()
